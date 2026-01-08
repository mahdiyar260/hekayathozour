from django.contrib import admin
from .models import Reserv, Phone, Days
from django.urls import path
from django.shortcuts import redirect
from django.utils.html import format_html
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side



@admin.register(Phone)
class PhoneAdmin(admin.ModelAdmin):
    list_display = ('id', 'phone_number')  # نمایش شماره و شناسه
    search_fields = ('phone_number',)      # امکان جستجو بر اساس شماره
    list_per_page = 20

@admin.register(Reserv)
class ReservAdmin(admin.ModelAdmin):
    list_display = ('leader_name', 'day', 'reservation_code', 'phone_number', 'men', 'women')
    list_filter = ('day',)          # فیلتر بر اساس روز
    search_fields = ('leader_name', 'reservation_code', 'phone_number')  # امکان جستجو

@admin.register(Days)
class DaysAdmin(admin.ModelAdmin):
    list_display = ('day',
                    'status',
                    'total_men',
                    'total_women',
                    'total_people',
                    'export_excel_button',
                    'toggle_status_button')

    def total_men(self, obj):
        total = Reserv.objects.filter(day=obj.day).aggregate(total=Sum('men'))['total'] or 0
        return total
    total_men.short_description = 'مردان'

    def total_women(self, obj):
        total = Reserv.objects.filter(day=obj.day).aggregate(total=Sum('women'))['total'] or 0
        return total
    total_women.short_description = 'زنان'

    def total_people(self, obj):
        stats = Reserv.objects.filter(day=obj.day).aggregate(
            men_sum=Sum('men'),
            women_sum=Sum('women')
        )
        return (stats['men_sum'] or 0) + (stats['women_sum'] or 0)
    total_people.short_description = 'کل افراد'

    def export_excel_button(self, obj):
        return format_html(
            '<a class="button" href="export-excel/{}/">دانلود اکسل</a>',
            obj.id
        )
    export_excel_button.short_description = 'دانلود اکسل'

    # دکمه تغییر وضعیت

    def toggle_status_button(self, obj):
        return format_html( '<a class="button" href="toggle-status/{}/"> تغییر وضعیت </a>', obj.id )
    toggle_status_button.short_description = 'تغییر وضعیت'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'toggle-status/<int:day_id>/',
                self.admin_site.admin_view(self.toggle_status),
                name='toggle-day-status'
            ),
            path(
                'export-excel/<int:day_id>/',
                self.admin_site.admin_view(self.export_excel),
                name='export-day-excel'
            ),
        ]
        return custom_urls + urls

    
    def toggle_status(self, request, day_id):
        day = Days.objects.get(id = day_id)
        if day.status == 'able':
            day.status = 'unable'
        else:
            day.status = 'able'
        day.save()
        return redirect(request.META.get('HTTP_REFERER'))

    #دانلود فایل اکسل

    def export_reservs_to_excel(self, day_obj):

        wb = Workbook()
        ws = wb.active
        ws.title = f"Day {day_obj.day.replace('/', '-')}"

        # ✅ راست به چپ
        ws.sheet_view.rightToLeft = True

        reservs = Reserv.objects.filter(day=day_obj.day)

        total_men = sum(r.men for r in reservs)
        total_women = sum(r.women for r in reservs)
        total_people = total_men + total_women

        # 🎨 استایل‌ها
        default_font = Font(name='Calibri', size=18)
        bold_font = Font(name='Calibri', size=18, bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        thick_border = Border(bottom=Side(style='thick'))

        # 🔹 خلاصه بالا
        summary_rows = [
            ('روز', day_obj.day),
            ('مجموع مردان', total_men),
            ('مجموع زنان', total_women),
            ('مجموع کل', total_people),
        ]
        for row in summary_rows:
            ws.append(row)
            for col in range(1, 3):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = bold_font
                cell.alignment = center_align

        # 🔻 خط ضخیم جداکننده
        ws.append([])
        sep_row = ws.max_row + 1
        for col in range(1, 7):
            ws.cell(row=sep_row, column=col).border = thick_border

        ws.append([])

        # 🔹 هدر جدول
        headers = ['نام سرگروه', 'کد رزرو', 'شماره تلفن', 'تعداد مردان', 'تعداد زنان', 'جمع کل']
        ws.append(headers)
        header_row = ws.max_row
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = bold_font
            cell.alignment = center_align

        # 🔹 داده‌ها
        for r in reservs:
            ws.append([r.leader_name, r.reservation_code, r.phone_number, r.men, r.women, r.men + r.women])
            for col in range(1, 7):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = default_font
                cell.alignment = center_align

        # 🔹 AutoFit شبیه Excel: یکسان کردن تمام ستون‌ها بر اساس طول بزرگترین متن
        max_length = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        auto_width = max_length + 4  # کمی فاصله برای زیبایی
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = auto_width

        # پاسخ دانلود
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"reservations_day_{day_obj.day}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response



    
    def export_excel(self, request, day_id):
        day = Days.objects.get(id=day_id)
        return self.export_reservs_to_excel(day)


    


# Register your models here.
